import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import workbook

#setando a aparencia padrao
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()

    def layout_config(self):
        self.title("Sistema de cadastro de produtos")
        self.geometry("700x500")
        self.resizable(False, False)

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text='Tema', bg_color='transparent', text_color=["#000", "#fff"])
        self.lb_apm.place(x=15, y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System default"], command=self.change_apm)
        self.opt_apm.place(x=15, y=460)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema de cadastro de produtos", font=("Century Gothic bold", 24), text_color="#fff")
        title.place(x=150, y=10)
        span = ctk.CTkLabel(self, text="Por favor, preencha todos os campos do formulario", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        span.place(x=50, y=70)

        ficheiro = pathlib.Path("Predinho.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro = openpyxl.Workbook()
            folha = ficheiro.active
            folha["A1"] = "Nome do produto"
            folha["B1"] = "Tipo de produto"
            folha["C1"] = "Data da compra"
            folha["D1"] = "Quem comprou"
            folha["E1"] = "Loja em que foi comprado"
            folha["F1"] = "Observacoes"

            ficheiro.save("Predinho.xlsx")

        def submit():
            
            #pegando os dados dos entrys
            name = name_value.get()  
            contact = contact_entry.get()
            age = age_value.get()          
            gender = gender_combobox.get()
            address = address_value.get()
            obs = obs_entry.get(0.0, END)

            ficheiro = openpyxl.load_workbook("Predinho.xlsx")
            folha = ficheiro.active
            folha.cell(column=1, row=folha.max_row+1, value = name)
            folha.cell(column=2, row=folha.max_row, value = contact)
            folha.cell(column=3, row=folha.max_row, value = age)
            folha.cell(column=4, row=folha.max_row, value = gender)
            folha.cell(column=5, row=folha.max_row, value = address)
            folha.cell(column=6, row=folha.max_row, value = obs)

            ficheiro.save(r"Predinho.xlsx")
            messagebox.showinfo(title="Sistema", message="Dados salvos com sucesso!")

        def clear():
            
            #limpando os dados dos entrys
            name_value.set("")
            age_value.set("")
            address_value.set("")
            obs_entry.delete(0.0, END)

#--------------------------------------------------------------------------------------------------------
        #text variables
        name_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()
        
        #Entrys

        #nome
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic bold", 16), fg_color="transparent")
        name_entry.place(x=50, y=150)

        #contato
        contact_entry = ctk.CTkComboBox(self, values=["Material de construcao", "Alvenaria", "Piso", "Vidro", "Tinta"], font=("Century Gothic bold", 14), width=150)
        contact_entry.set("")
        contact_entry.place(x=450, y=150)

        #idade
        age_entry = ctk.CTkEntry(self, width=150, textvariable=age_value, font=("Century Gothic bold", 16), fg_color="transparent")
        age_entry.place(x=300, y=220)

        #genero
        gender_combobox = ctk.CTkComboBox(self, values=["Janice", "Cristiane", "Carlos", "Luis Afonso", "Marcos"], font=("Century Gothic bold", 14), width=150)
        gender_combobox.set("")
        gender_combobox.place(x=500, y=220)

        #endereco
        address_entry = ctk.CTkEntry(self, width=230, textvariable=address_value, font=("Century Gothic bold", 16), fg_color="transparent")
        address_entry.place(x=50, y=220)

        #Entrada de observacoes
        obs_entry = ctk.CTkTextbox(self, width=470, height=150, font=("arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")
        obs_entry.place(x=180, y=260)

#--------------------------------------------------------------------------------------------------------
        #Labels
        
        #nome
        lb_name = ctk.CTkLabel(self, text="Nome do produto", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_name.place(x=50, y=120)
        
        #contato
        lb_contact = ctk.CTkLabel(self, text="Tipo de produto", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_contact.place(x=450, y=120)
       
        #idade
        lb_age = ctk.CTkLabel(self, text="Data da compra", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_age.place(x=300, y=190)
        
        #genero
        lb_gender = ctk.CTkLabel(self, text="Quem comprou", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_gender.place(x=500, y=190)
       
        #endereco
        lb_address = ctk.CTkLabel(self, text="Loja em que foi comprado", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_address.place(x=50, y=190)
       
        #observacao
        lb_obs = ctk.CTkLabel(self, text="Observacoes", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_obs.place(x=50, y=260)
        
#--------------------------------------------------------------------------------------------------------
        #botoes
        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131")
        btn_submit.place(x=300, y=420)

        btn_clear = ctk.CTkButton(self, text="Limpar campos".upper(), command=clear, fg_color="#555", hover_color="#333")
        btn_clear.place(x=500, y=420)

    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)


if __name__ == "__main__":
    app = App()
    app.mainloop()